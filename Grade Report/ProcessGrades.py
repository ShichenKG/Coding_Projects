#Robert Ball
#Process Grades
import pandas as pd
import smtplib, ssl
from email import encoders
from email.mime.base import MIMEBase
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from openpyxl import Workbook
import PySimpleGUI as sg

sender_email = "softwarejrti@gmail.com"
password = "Coding63"


smtp_server = "smtp.gmail.com"
port = 587
context = ssl.create_default_context()

teacherList = ['Albright','Arntz','Bennett','Brown','Butcher','Christman','Cunningham','Eisenhart','Files','Gerda',
                'Harris', 'Jlockhart','KBall','lantz','LAYHEW','Morgan','Odom','Price','RBall','Rlockhart','SULLIVAN',
                'Ware','Zeger','Vanorsdale']

teacherEmail = {'ALBRIGHT' : 'kyle.albright@k12.wv.us','ARNTZ' : 'carntz@k12.wv.us','BENNETT' : 'wybennet@k12.wv.us',
                    'BROWN' : 'sjbrown@k12.wv.us','BUTCHER' : 'jlbutche@k12.wv.us','CHRISTMAN' : 'vchristm@k12.wv.us' ,
                    'CUNNINGHAM' : 'lauren.cunningham@k12.wv.us','EISENHART': 'paul.eisenhart@k12.wv.us',
                    'FILES' : 'afiles@k12.wv.us','GERDA' : 'dgerda@k12.wv.us', 'HARRIS' : 'steven.harris@k12.wv.us',
                    'JLOCKHART' : 'jlockhart@k12.wv.us','KBALL' : 'kenda.ball@k12.wv.us','LANTZ' : 'llantz@k12.wv.us',
                    'LAYHEW' : 'llayhew@k12.wv.us','MORGAN' : 'm.d.morgan@k12.wv.us','ODOM' :'rodom@k12.wv.us',
                    'PRICE' : 'kwprice@k12.wv.us','RBALL': 'Robert.w.ball@k12.wv.us','RLOCKHART' : 'rebecca.lockhart@k12.wv.us',
                    'SULLIVAN' : 'mrsulliv@k12.wv.us', 'WARE' : 'jeware@k12.wv.us', 'ZEGER' : 'bzeger@k12.wv.us',
                    'VANORSDALE' : 'jvanorsd@k12.wv.us'}


def creatSpreadsheet():
    pass

# Send email
def send(address, body, subject, filename, file):
    Email = MIMEMultipart()
    Email["To"] = address
    Email["Subject"] = subject
    Email.attach(MIMEText(body, "plain"))
    print(Address)

    try:
        # Attach file,
        if file is not None:
            with open(file, "rb") as File:
                Attachment = MIMEBase("application", "octet-stream")
                Attachment.set_payload(File.read())
                if filename is None:
                    filename = File.name
            # Encode file so it can be sent through,
            encoders.encode_base64(Attachment)

            Attachment.add_header(
                "Content-Disposition",
                "attachment; filename=" + filename,
            )

            Email.attach(Attachment) # Attachment

        # Setup server stuff
        server = smtplib.SMTP(smtp_server, port)
        server.starttls(context=context)

        server.login(sender_email, password) # Login

        #server.sendmail(sender_email, address, Email.as_string())  # Send email
        #server.sendmail(sender_email, "robert.w.ball@k12.wv.us", Email.as_string())
        #server.sendmail((sender_email, "asalbrig@k12.wv.us", Email.as_string()))


    # Return errors
    except Exception as e:
        return e

    else:
        return True # Return True if all goes well

#######
countteach = 1
def getSchoolName(schoolCode):
    Musslemen = "MUHS"
    Washington = "WAHS"
    Jefferson = "JEHS"
    Martinsburg = "MAHS"
    Springmills = "SMHS"
    BerkeleySprings = "BSHS"
    PawPaw = "PPHS"
    Hedgesville = "HEHS"
    HomeSchool = "HOME"

    if schoolCode == Musslemen:
        return "Musslemen"
    elif schoolCode == Washington:
        return "Washington"
    elif schoolCode == Jefferson:
        return "Jefferson"
    elif schoolCode == Martinsburg:
        return "Martinsburg"
    elif schoolCode == Springmills:
        return "Springmills"
    elif schoolCode == BerkeleySprings:
        return "Berkeley Springs"
    elif schoolCode == PawPaw:
        return "Paw Paw"
    elif schoolCode == Hedgesville:
        return "Hedgesville"
    elif schoolCode == HomeSchool:
        return "Home Schooled"
    else:
        return "Faculty"

##################################
# Edit by: Shane M.D.
# Senior, PM
# Window wrap and progress bar
##################################
sg.LOOK_AND_FEEL_TABLE['JRTI'] = {'BACKGROUND': '#2E75B6',
                                        'TEXT': 'black',
                                        'INPUT': '#FFFFFF',
                                        'TEXT_INPUT': '#000000',
                                        'SCROLL': '#41719C',
                                        'BUTTON': ('black', '#41719C'),
                                        'PROGRESS': ('#003058', '#FFFFFF'),
                                        'BORDER': 1, 'SLIDER_DEPTH': 0, 'PROGRESS_DEPTH': 1,
                                        }
sg.theme('JRTI')
num = 1
font = ['Times New Roman', 18]
font2 = ['Times New Roman', 14]
Term = ['Term 1','Term 2','Term 3','Term 4','Semester 1','Semester 2','Year']
layout = [[sg.Text('Powered by:',relief='sunken',font=(font)),
           sg.Image(r'D:\Shane\Python Projects\Grade Report\Logo.png')],
          [sg.Text('Teachers:       ',font=(font)),
           sg.Combo(teacherList, size=(30,30),enable_events=True, key='-Teach-',font=(font2),readonly=True),
           sg.Checkbox('All Teachers',enable_events=True, key='-CheckTeach-',font=(font), default=True)],
          [sg.Text('School Term: ',font=(font)),
           sg.Combo(Term, size=(30,30),enable_events=True, key='-Term-',font=(font2),readonly=True)],
          [sg.ProgressBar(max_value=24, orientation='h', size=(40, 30), key='-Progress-'),
           sg.Text('Waiting...',enable_events=True, key='-Text-',font=(font))],
          [sg.InputText(size=(36, 2),key='-Entry-',font=(font),readonly=True),
           sg.Button('Send',key='-Submit-',size=(6,1),font=(font2),bind_return_key=True),
           # Having an issue with specific file browsing
           sg.FileBrowse(key='-File-',size=(6,1),font=(font2))],
          [sg.StatusBar("This is the status bar",key='-Stat-',font=(font))]]
window = sg.Window('Grade Report v1.3', layout, size=(630,300),finalize=True)

# Main loop where events and 
while True:
    event, values = window.read()
    print(event,values)
    if event == sg.WINDOW_CLOSED or event == 'Quit':
        break
    elif event == '-Submit-':
        data = pd.read_csv('grades1.csv', encoding='cp1252')
        data.sort_values(by='Course Name')
        Body = "This email was automatically created by JRTI Software!\n \n" \
               "Attached are the interim grades for all your students from Schoology as of the date of this email.\n" \
               "This is an early run to allow you to examine the data for any discrepancies.\n" \
               "To identify your classes, it is critical that you added your name to the section name in Schoology\n" \
               "If your file is empty it means the program did not find any grades associated with your name. \n" \
               "These reports will be regenerated as soon as the deadline for interim grades has passed.\n\n" \
               "If you have any questions: Please contact Bob Ball - Robert.w.ball@k12.wv.us"

        # Goes through list of teachers and creates files for them
        for Teacher in teacherList:
            destinationFileName = f"InterimGrades-{Teacher}.xlsx"
            wb = Workbook()
            ws = wb.active
            ws.title = "Interim Grades"
            rowCount = 1

            # Iterates through data and loads information
            for index, row in data.iterrows():
                userId = row["Unique User ID"]
                schoolCode = userId[:4]
                schoolName = getSchoolName(schoolCode)
                firstName = row["First Name"]
                lastName = row["Last Name"]
                courseName = row["Course Name"]
                sectionName = row["Section Name"]
                gradingPeriod = row['Grading Period']
                markingPeriod = row['Titles']
                grade = row['Grades']
                letterGrade = row["Letter Grades"]

                # Eleminates faculty and makes sure it's marking period 1, and loads the cells
                if schoolName != "Faculty" and markingPeriod == '*Mkp 1':
                    printRow = f"{lastName},{firstName} {schoolName} \t {courseName} \t {sectionName} \t {gradingPeriod} \t {grade} \t {letterGrade} \n"
                    if Teacher.upper() in (sectionName.upper()):
                        ws.cell(rowCount, column=1, value=userId)
                        ws.cell(rowCount, column=2, value=schoolCode)
                        ws.cell(rowCount, column=3, value=schoolName)
                        ws.cell(rowCount, column=4, value=firstName)
                        ws.cell(rowCount, column=5, value=lastName)
                        ws.cell(rowCount, column=6, value=courseName)
                        ws.cell(rowCount, column=7, value=sectionName)
                        ws.cell(rowCount, column=8, value=gradingPeriod)
                        ws.cell(rowCount, column=9, value=markingPeriod)
                        ws.cell(rowCount, column=10, value=grade)
                        ws.cell(rowCount, column=11, value=letterGrade)
                        rowCount += 1

            wb.save(destinationFileName)
            wb.close()
            Address = teacherEmail[Teacher.upper()]
            # Updates once every loop creating a "..." effect
            if num == 1:
                num += 1
                window['-Text-'].update('Processing.')
            elif num == 2:
                num += 1
                window['-Text-'].update('Processing..')
            elif num == 3:
                num = 1
                window['-Text-'].update('Processing...')

            # This counts each time the for loop runs and adds a counter with the teacher email
            countteach += 1
            window['-Progress-'].update(countteach)
            result = send(Address, Body, "Interim Grade Reports", destinationFileName, destinationFileName)
            window['-Stat-'].update(Address)
            if countteach >= len(teacherList):
                window['-Stat-'].update('Reports sent!')
                window['-Text-'].update('Finished!')

    else:
        print(event,values)
